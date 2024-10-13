
import pandas as pd
from datetime import datetime, timedelta
import csv
import time
import pickle
from selenium.webdriver.common.by import By
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import undetected_chromedriver as uc
import os
import csv
import tkinter as tk
from tkinter import simpledialog
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import re
# Import the necessary module
from bs4 import BeautifulSoup
import random

# Define the range for the random sleep time in seconds
min_sleep_time, max_sleep_time = 1, 3

# Generate and use a random sleep time within the defined range
random_sleep=time.sleep(random.uniform(min_sleep_time, max_sleep_time))



options = uc.ChromeOptions()
options.headless = False

driver = uc.Chrome(options=options, use_subprocess=True)
# Set the browser window size
driver.set_window_size(1000, 800)
####################################################################

def check_folder_empty(folder):
    if not os.path.exists(folder):
        os.makedirs(folder)
    return not any(filename.endswith('.csv') for filename in os.listdir(folder))

def split_csv(input_file, output_folder, max_records=500):
    if not check_folder_empty(output_folder):
        print("CSV files already exist in the output folder. Skipping splitting.")
        return

    with open(input_file, 'r', newline='') as csvfile:
        reader = csv.reader(csvfile)
        header = next(reader)  # Read the header

        count = 0
        file_index = 1
        os.makedirs(output_folder, exist_ok=True)
        output_file = os.path.join(output_folder, f'split_{file_index}.csv')
        with open(output_file, 'w', newline='') as outfile:
            writer = csv.writer(outfile)
            writer.writerow(header)  # Write the header to each split file
            for row in reader:
                writer.writerow(row)
                count += 1
                if count == max_records:
                    count = 0
                    file_index += 1
                    output_file = os.path.join(output_folder, f'split_{file_index}.csv')
                    outfile.close()  # Close the current file
                    outfile = open(output_file, 'w', newline='')  # Reopen the file
                    writer = csv.writer(outfile)
                    writer.writerow(header)

# Function to select a CSV file from the split folder
def select_csv_from_split_folder(output_folder):
    csv_files = [filename for filename in os.listdir(output_folder) if filename.endswith('.csv')]
    if not csv_files:
        print("No CSV files found in the split folder.")
        return None
    else:
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes("-topmost", 1)  # Bring the dialog window to the front
        options = "\n".join([f"{i+1}: {file}" for i, file in enumerate(csv_files)])
        message = f"Choose the index of the CSV file to work on:\n{options}"
        index = simpledialog.askinteger("Select CSV file", message, minvalue=1, maxvalue=len(csv_files))
        if index:
            return os.path.join(output_folder, csv_files[index - 1])
        else:
            return None

# Example usage:
input_file = 'links.csv'  # Replace 'input.csv' with your CSV file path
output_folder = 'split-csv'  # Folder where split CSV files will be saved
split_csv(input_file, output_folder, max_records=500)

# Select a CSV file from the split folder
csv_file = select_csv_from_split_folder(output_folder)
print("Selected CSV file:", csv_file)



##################################################

def login(driver, cookies_file_path='cookies.pkl'):
    if os.path.exists(cookies_file_path):
        # Load cookies if the file exists
        with open(cookies_file_path, 'rb') as cookies_file:
            cookies = pickle.load(cookies_file)
            # Use the loaded cookies for the session
            driver.get('https://www.youtube.com')
            driver.implicitly_wait(10)
            for cookie in cookies:
                driver.add_cookie(cookie)
            print("Cookies added")
            driver.refresh()
            driver.implicitly_wait(10)
    else:

        driver.get('https://www.youtube.com')
        driver.implicitly_wait(15)


        try:
            clk=driver.find_element(By.XPATH,"(//div[@class='yt-spec-touch-feedback-shape yt-spec-touch-feedback-shape--touch-response-inverse'])[2]").click()
            time.sleep(random.uniform(1, 3))
            driver.implicitly_wait(10)
        except:
            pass
        try:
            driver.find_element(By.XPATH,"//*[@id='buttons']/ytd-button-renderer/yt-button-shape/a/yt-touch-feedback-shape/div").click()
            time.sleep(random.uniform(1, 3))
            driver.implicitly_wait(10)
            element=driver.find_element(By.XPATH,"//*[@id='identifierId']").click()
            print("cLICKED")
            # Find and input the email/username
            email_field = driver.find_element(By.XPATH, "//*[@id='identifierId']")
            email_field.send_keys("darkjhon449")  # Replace with your email/username
            time.sleep(random.uniform(1, 3))  # Adding a delay to ensure the text is inputted before continuing

            # Press Enter to proceed to the password input
            email_field.send_keys(Keys.ENTER)
            time.sleep(random.uniform(1, 3))  # Adding a delay to ensure the page loads properly
            driver.implicitly_wait(10)

            email_pass = driver.find_element(By.XPATH, "//*[@id='password']/div[1]/div/div[1]/input")
            email_pass.send_keys("Pakistan@1947")  # Replace with your email/username
            time.sleep(1)  # Adding a delay to ensure the text is inputted before continuing

            # Press Enter to proceed to the password input
            email_pass.send_keys(Keys.ENTER)
            time.sleep(random.uniform(1, 3)) # Adding a delay to ensure the page loads properly
            driver.implicitly_wait(10)
            input()



        except:
            pass





        # Save cookies for future use
        with open(cookies_file_path, 'wb') as cookies_file:
            pickle.dump(driver.get_cookies(), cookies_file)
        print("cookies saved")

login(driver=driver)

##################################

def create_excel_file():
    # Create a folder named "successful_Data" if it doesn't exist
    folder_name = "successful_Data"
    folder_name2 = "failed_Data"

    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
        print(f"Folder '{folder_name}' created successfully.")

    if not os.path.exists(folder_name2):
        os.makedirs(folder_name2)
        print(f"Folder '{folder_name2}' created successfully.")

    # Get the current date in the desired format
    current_date = datetime.now().strftime("-%d-%m-%Y")
    file_name = os.path.join(folder_name, f"data{current_date}.xlsx")

    # Check if a file with the current date exists
    if not os.path.exists(file_name):
        # If it doesn't exist, create a new Excel file with that name
        pd.DataFrame().to_excel(file_name, index=False)
        print(f"Excel file '{file_name}' created successfully.")
    else:
        print(f"Excel file '{file_name}' already exists.")

    # Create Excel files in "failed_Data" folder
    failed_folder_name = "failed_Data"
    Banned_or_no_short_video = os.path.join(failed_folder_name, "No_Short_Videos_Or_Banned.xlsx")
    no_video_in_date_range = os.path.join(failed_folder_name, f"No_Short_Video_in_date-Range{current_date}.xlsx")

    # Check if the files already exist
    if not os.path.exists(Banned_or_no_short_video):
        # Create the file only if it doesn't exist
        pd.DataFrame().to_excel(Banned_or_no_short_video, index=False)
        print(f"Excel file '{Banned_or_no_short_video}' created successfully.")
    else:
        print(f"Excel file '{Banned_or_no_short_video}' already exists.")

    if not os.path.exists(no_video_in_date_range):
        # Create the file only if it doesn't exist
        pd.DataFrame().to_excel(no_video_in_date_range, index=False)
        print(f"Excel file '{no_video_in_date_range}' created successfully.")
    else:
        print(f"Excel file '{no_video_in_date_range}' already exists.")

    return file_name, Banned_or_no_short_video, no_video_in_date_range

excel_file, Banned_or_no_short_video, no_video_in_date_range = create_excel_file()


###################################
headers = ["Channel Name", "youtube url", "subscribers", "Check Date", "Released Date", "posted in date range",
               "Shorts Name", "Description", "Likes", "Views", "link"]

df = pd.DataFrame(columns=headers)
check_date = datetime.now().strftime("%d-%m-%Y")
############################################

while True:
    try:
        # Read channel URLs from the CSV file
        with open(csv_file, "r") as file:
            reader = csv.DictReader(file)
            rows = list(reader)  # Convert the reader object to a list of rows for easier access
            for index, row in enumerate(rows):
                check_date = datetime.now().strftime("%d-%m-%Y")
                url = row['links']
                processed = row['processed']
                if not processed:  # Process only if 'processed' column is empty
                    # print(url)
                    chanel_url = url  # Store the URL in chanel_url variable
                    print(f"Processing link  ##  " + url)

                    print("loading url")
                    # Load the channel URL in the browser
                    driver.get(url.strip())  # Strip any leading/trailing whitespaces
                    driver.implicitly_wait(10)
                    fixed_url = driver.current_url
                    # Get page source



                    # Update the processed column for the current URL to 'Yes'
                    rows[index]['processed'] = 'Yes'

                    # Write back to the CSV file with updated data
                    with open(csv_file, "w", newline='') as csvfile:
                        fieldnames = ['links', 'processed']
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writeheader()
                        writer.writerows(rows)
                     # Break the loop after processing the current URL

                    page_source = driver.page_source

                    # Parse page source using BeautifulSoup
                    soup = BeautifulSoup(page_source, 'html.parser')

                    # Extract all text from the page
                    all_text = soup.get_text()


                    if (
                            " terminated because we received multiple third-party claims of copyright" in all_text or "This channel does not exist" in all_text or "The page you were looking for doesn't exist" in all_text or
                            "account has been terminated because it is linked to an account that received multiple third-party claims of copyright infringement" in all_text) or "channel was removed because it violated our Community Guidelines" in all_text:

                        print("banned account")
                        #saving data
                        new_row = {'youtube url': fixed_url,
                                   'Check Date': check_date,
                                   'posted in date range': "Banned"}

                        ##################################################################################################################################
                        print("Trying 1")
                        ############################################
                        # Load the existing workbook
                        try:
                            wb = load_workbook(Banned_or_no_short_video)
                            ws = wb.active
                        except FileNotFoundError:
                            # If the file doesn't exist, create a new workbook
                            wb = Workbook()
                            ws = wb.active

                        # Write data
                        if not ws['A1'].value:  # If the first cell is empty, assume headers are not written
                            headers = list(new_row.keys())
                            for col, header in enumerate(headers, start=1):
                                ws.cell(row=1, column=col, value=header)

                        # Find the next available row
                        row_num = ws.max_row + 1

                        # Write data to the next available row
                        for col, value in enumerate(new_row.values(), start=1):
                            ws.cell(row=row_num, column=col, value=value)

                        # Save the workbook
                        wb.save(Banned_or_no_short_video)

                        print("data is saved")




                        break
                    ##############################################
                    # Get the page source
                    page_source = driver.page_source

                    # Parse the page source using BeautifulSoup
                    soup = BeautifulSoup(page_source, 'html.parser')

                    # Find the element matching the XPath
                    element = soup.find("yt-tab-shape", {"tab-title": "Shorts"})

                    #######################################
                    find_short = None
                    short_xpath = "//*[@id='tabsContent']/yt-tab-group-shape/div[1]/yt-tab-shape[{}]"

                    for i in range(1, 4):  # Try different indexes from 1 to 9
                        try:
                            short_xpath_formatted = short_xpath.format(i)
                            find_short = driver.find_element(By.XPATH, short_xpath_formatted).text
                            if "Shorts" in find_short:
                                print(f"Found 'Shorts' at XPath: {short_xpath_formatted}")
                                click_short = driver.find_element(By.XPATH, short_xpath_formatted).click()
                                break
                        except:
                            pass

                    if find_short is None:
                        print("Text 'Shorts' not found in any of the XPaths.")
                        # saving data
                        new_row = {'youtube url': fixed_url,
                                   'Check Date': check_date,
                                   'posted in date range': "No short Videos"}

                        ##################################################################################################################################
                        print("Trying 2")
                        ############################################
                        # Load the existing workbook
                        try:
                            wb = load_workbook(Banned_or_no_short_video)
                            ws = wb.active
                        except FileNotFoundError:
                            # If the file doesn't exist, create a new workbook
                            wb = Workbook()
                            ws = wb.active

                        # Write data
                        if not ws['A1'].value:  # If the first cell is empty, assume headers are not written
                            headers = list(new_row.keys())
                            for col, header in enumerate(headers, start=1):
                                ws.cell(row=1, column=col, value=header)

                        # Find the next available row
                        row_num = ws.max_row + 1

                        # Write data to the next available row
                        for col, value in enumerate(new_row.values(), start=1):
                            ws.cell(row=row_num, column=col, value=value)

                        # Save the workbook
                        wb.save(Banned_or_no_short_video)

                        print("data is saved")

                        continue

                    else:
                        print(find_short)

                    # Continue with your code using the found XPath (short_xpath_formatted)

                    print(chanel_url)
                    try:
                        chhanel_name = driver.find_element(By.XPATH,
                                                           "(//*[@id='text'])[1]").text


                        #chhanel_name = driver.find_element(By.XPATH, "(//div[@class='style-scope ytd-channel-name'])[1]").text

                        print("Channel Name:", chhanel_name)
                    except:
                        chhanel_name=""




                    try:

                        subscribers = driver.find_element(By.XPATH, "//*[@id='subscriber-count']").text
                        print(subscribers)

                        #print(chhanel_name)
                    except :
                        try:
                            text1 = driver.find_element(By.XPATH, "//yt-content-metadata-view-model[@class]").text
                            print(subscribers)

                            # Define patterns using regular expressions
                            subscribers_pattern = r'(\d+)K subscribers'


                            # Extract subscribers count using regex
                            subscribers_match = re.search(subscribers_pattern, text1)
                            if subscribers_match:
                                subscribers_count = subscribers_match.group(1)
                            else:
                                subscribers_count = None



                            print("Subscribers:", subscribers_count)


                            subscribers = subscribers_count

                        except:
                            subscribers=""




                    ############


                    shorts = driver.find_elements(By.XPATH,
                                                  "//a[@class='yt-simple-endpoint focus-on-expand style-scope ytd-rich-grid-slim-media']")
                    print(len(shorts))

                    if shorts is None or len(shorts) < 1:

                        print("No Short Videos loop breaked")
                        # saving data
                        new_row = {'youtube url': fixed_url,
                                   'Check Date': check_date,
                                   'posted in date range': "No short Videos"}

                        ##################################################################################################################################
                        print("Trying 3")
                        ############################################
                        # Load the existing workbook
                        try:
                            wb = load_workbook(Banned_or_no_short_video)
                            ws = wb.active
                        except FileNotFoundError:
                            # If the file doesn't exist, create a new workbook
                            wb = Workbook()
                            ws = wb.active

                        # Write data
                        if not ws['A1'].value:  # If the first cell is empty, assume headers are not written
                            headers = list(new_row.keys())
                            for col, header in enumerate(headers, start=1):
                                ws.cell(row=1, column=col, value=header)

                        # Find the next available row
                        row_num = ws.max_row + 1

                        # Write data to the next available row
                        for col, value in enumerate(new_row.values(), start=1):
                            ws.cell(row=row_num, column=col, value=value)

                        # Save the workbook
                        wb.save(Banned_or_no_short_video)

                        print("data is saved")

                        print(
                            "33333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333")
                        break

                    # Store links in a list
                    links = []

                    for short in shorts:
                        link = short.get_attribute("href")
                        links.append(link)
                        # print(link)

                    print("all links saved")
                    # Set a flag to keep track of whether the 'if' part of the loop has executed
                    if_part_executed = False
                    # Iterate over the links and load them one by one
                    for link in links:
                        print("Loading link:", link)
                        driver.get(link)
                        driver.implicitly_wait(10)
                        try:

                            short_name = driver.find_element(By.XPATH,
                                                             "//h2[@class='title style-scope reel-player-header-renderer']").text
                        except:
                            continue

                        print(short_name)

                        time.sleep(random.uniform(1, 3))

                        print("click")
                        clk = driver.find_element(By.XPATH,
                                                  "//*[@id='button-shape']/button/yt-touch-feedback-shape/div/div[2]").click()
                        time.sleep(random.uniform(1, 3))
                        clk2 = driver.find_element(By.XPATH,
                                                   "//*[@id='items']/ytd-menu-service-item-renderer/tp-yt-paper-item/yt-formatted-string").click()
                        print("clicked")
                        try:
                            description = driver.find_element(By.XPATH, "//div[@id='description']").text
                        except:
                            description = ""
                        print(description)
                        try:

                            likes1 = driver.find_element(By.XPATH,
                                                         "(//div[@class='YtwFactoidRendererFactoid'])[1]").text
                        except:
                            likes1 = "Likes"
                        print(likes1)
                        likes = likes1.replace('Likes', '')
                        try:

                            veiws1 = driver.find_element(By.XPATH,
                                                         "(//div[@class='YtwFactoidRendererFactoid'])[2]").text
                        except:
                            veiws1 = "Views"
                        print(veiws1)
                        veiws = veiws1.replace('Views', '')
                        try:

                            date = driver.find_element(By.XPATH, "(//div[@class='YtwFactoidRendererFactoid'])[3]").text
                        except:
                            date = "\n"

                        date_string = date
                        date = date.replace('\n', ' ')
                        print(date)
                        # Date string from the webpage
                        date_str = date
                        # Convert to datetime object
                        try:
                            # Try parsing the date assuming it's in the format "Apr 26 2024"
                            date_obj = datetime.strptime(date_str, "%b %d %Y")
                        except ValueError:
                            # If ValueError occurs, try parsing the date assuming it's in the format "26 Apr 2024"
                            date_obj = datetime.strptime(date_str, "%d %b %Y")

                        # Format the date consistently
                        date = date_obj.strftime("%b %d %Y")
                        print(date)


                        # Get the current date
                        current_date = datetime.now()

                        # Generate a list of date strings for the past 7 days
                        date_range = [(current_date - timedelta(days=i)).strftime("%b %d %Y") for i in range(7)]
                        # Print the generated date range
                        # print("Date Range (Past 7 days):")

                        for dates in date_range:
                            print(dates)
                        #input()

                        # Check if the date falls within the past 7 days
                        if (date in date_range) or ("ago" in date.lower()):

                            if_part_executed = True



                            new_row = {'Channel Name': chhanel_name, 'youtube url': fixed_url,
                                       'subscribers': subscribers,
                                       'Check Date': check_date, 'Released Date': date,
                                       'posted in date range': "Yes", 'Shorts Name': short_name,
                                       'Description': description, 'Likes': likes, 'Views': veiws,
                                       'link': link}

                            ##################################################################################################################################
                            print("Trying 4")
                            ############################################
                            # Load the existing workbook
                            try:
                                wb = load_workbook(excel_file)
                                ws = wb.active
                            except FileNotFoundError:
                                # If the file doesn't exist, create a new workbook
                                wb = Workbook()
                                ws = wb.active

                            # Write data
                            if not ws['A1'].value:  # If the first cell is empty, assume headers are not written
                                headers = list(new_row.keys())
                                for col, header in enumerate(headers, start=1):
                                    ws.cell(row=1, column=col, value=header)

                            # Find the next available row
                            row_num = ws.max_row + 1

                            # Write data to the next available row
                            for col, value in enumerate(new_row.values(), start=1):
                                ws.cell(row=row_num, column=col, value=value)

                            # Save the workbook
                            wb.save(excel_file)

                            print("data is saved")

                            print("loop is continueing")


                            print("%" * 100)
                            print(
                                "4444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444")

                        else:
                            if not if_part_executed:  # If the 'if' part hasn't executed in the first iteration

                                check_date = datetime.now().strftime("%d-%m-%Y")
                                new_row = {'Channel Name': chhanel_name, 'youtube url': fixed_url,
                                           'subscribers': subscribers,
                                           'Check Date': check_date, 'Released Date': date,
                                           'posted in date range': "No",
                                           'Shorts Name': short_name,
                                           'Description': description,
                                           'Likes': likes, 'Views': veiws, 'link': link}

                                print("Trying 5")
                                ############################################
                                # Load the existing workbook
                                try:
                                    wb = load_workbook(no_video_in_date_range)
                                    ws = wb.active
                                except FileNotFoundError:
                                    # If the file doesn't exist, create a new workbook
                                    wb = Workbook()
                                    ws = wb.active

                                # Write data
                                if not ws['A1'].value:  # If the first cell is empty, assume headers are not written
                                    headers = list(new_row.keys())
                                    for col, header in enumerate(headers, start=1):
                                        ws.cell(row=1, column=col, value=header)

                                # Find the next available row
                                row_num = ws.max_row + 1

                                # Write data to the next available row
                                for col, value in enumerate(new_row.values(), start=1):
                                    ws.cell(row=row_num, column=col, value=value)

                                # Save the workbook
                                wb.save(no_video_in_date_range)
                                print("data is saved")


                                print("loop breaked")

                                print(
                                    "555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555")

                                break
                            if (date not in date_range) or ("ago"  not in date.lower()):
                                print("breaking loop after saving data")

                                break




    except Exception as e:
        #print(e)

        pass








